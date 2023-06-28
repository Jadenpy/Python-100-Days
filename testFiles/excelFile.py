# -*- coding: utf-8 -*-
# @Author: Jaden Hu
# @Date:   2023-06-28 14:09:22
# @Last Modified by:   Jaden Hu
# @Last Modified time: 2023-06-28 14:10:21


import datetime

from openpyxl import Workbook


wb = Workbook()
ws = wb.active

ws['A1'] = 42
ws.append([1, 2, 3])
ws['A2'] = datetime.datetime.now()

wb.save("sample.xlsx")

